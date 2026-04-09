import requests
from datetime import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

shops = [
 {
        "name": "0711Card",
        "url": "https://0711card.de",
        "ck": "ck_877cb4e0a1cf1f39aae5381799886eef0e22d621",
        "cs": "cs_41712015c9e3efb61bc92e76ab16459582910f64"
    },  
    {
        "name": "AbensbergCard",
        "url": "https://abensbergcard.de",
        "ck": "ck_dd359163a1866893bb4d0a941b8d566ba70b7f4f",
        "cs": "cs_d74a775c32c2e3c5835086ff65b0e13ade2a874b"
    },
    {
        "name": "Baden-Baden-Card",
        "url": "https://baden-baden-card.de",
        "ck": "ck_613756ccddf6da3430ba25a0a9414b908a230f89",
        "cs": "cs_4ca493610f4e74653ba7bf5c7eb1e7f7e9c75230"
    },
    {
        "name": "Bad Waldsee CityCard",
        "url": "https://badwaldseecitycard.de",
        "ck": "ck_7f85c7615a91adc05ef8366f2e23fb70a64560e8",
        "cs": "cs_8a8bf37328e52b524f4a023d1f54230002e76db2"
    },
    {
        "name": "BalingenCard",
        "url": "https://balingencard.de",
        "ck": "ck_a5cdad516ad49cd729676a6d3af7ff22889a5dc6",
        "cs": "cs_d417b6bb38cd594ce99e67c6634a0ec3b1c1505e"
    },
    {
        "name": "BesigheimCard",
        "url": "https://besigheimcard.de",
        "ck": "ck_ab4e1809b7f3188dcb3f3f6a36b1522cf9422ddd",
        "cs": "cs_3e708ae1d948760fa5fad0b40a2d8c5547f54ba4"
    },
    {
        "name": "SchwabachCard",
        "url": "https://schwabachcard.de",
        "ck": "ck_41fd8cd5919d8951e3b7f614c73cdc92136f0f55",
        "cs": "cs_107e6dfef7d7558d84a2057ebe2784610f547dd5"
    },
    {
        "name": "CalwCard",
        "url": "https://calwcard.de",
        "ck": "ck_f5c46e2f7e5670b696f88eeb1239bd6465debd74",
        "cs": "cs_fb590ef4e62610183a06a29fec27e1a64e6cfd0b"
    },
    {
        "name": "DETTINGEN ERMSCARD",
        "url": "https://dettingen-ermscard.de",
        "ck": "ck_80d540ae2f1a4bed6f5c3b7ce8b147519cabcc09",
        "cs": "cs_03754ccedb1403bdcfb818bad9e19342621484f0"
    },
    {
        "name": "ECHT FREIBURG CARD",
        "url": "https://echtfreiburgcard.de",
        "ck": "ck_4442abedf50a5217c4c40a331cf3433c0e797cb5",
        "cs": "cs_c2bfc2d2136ecd6d4d642f5aa3a35fe364912365"
    },
    {
        "name": "EttenheimCard",
        "url": "https://ettenheimcard.de",
        "ck": "ck_58e0ca631159bc8a16a951f687df1a09149a3361",
        "cs": "cs_8ade17026c6863d18c7bd34e140bc6cc191f4b44"
    },
    {
        "name": "Erlebnisregion Europa-Park-Card",
        "url": "https://erlebnisregion-europaparkcard.de",
        "ck": "ck_0968ea036a01f54dde55ec327993abcb09ceac77",
        "cs": "cs_459c8e36fcc22e0fcf3eef2e4f05e258b1d92c70"
    },
    {
        "name": "FördeCard",
        "url": "https://foerdecard.com",
        "ck": "ck_0b99008d57b5f5d6843ac1d7f30c15a8e4177f3f",
        "cs": "cs_b2f767dbe04cd0d9e6677c93cad67f0e5fe09535"
    },
    {
        "name": "HaslachCard",
        "url": "https://haslachcard.de",
        "ck": "ck_470754a7515c4eb93d2f5dfe14fdebd12939a7b5",
        "cs": "cs_3d3b24bd0e7dc1551e06e35d1488fd04b858e4c6"
    },
    {
        "name": "HerbolzheimCard",
        "url": "https://herbolzheim-karte.de",
        "ck": "ck_faf8f61d23fdceb4356c7dc0a23e72bd79036f0b",
        "cs": "cs_11944a48bfd3849920e38b6e89c16de9e3fd88d8"
    },
    {
        "name": "HU-SmartCard",
        "url": "https://hu-smartcard.de",
        "ck": "ck_bae9d287e70fcbf829959c74a6abd14bbcf683d3",
        "cs": "cs_873c15f62f0141d239b5841f876625aff482d389"
    },
    {
        "name": "KenzingenCard",
        "url": "https://kenzingencard.de",
        "ck": "ck_27561320d969cc4f05d67bee5535ef01dd78238d",
        "cs": "cs_6576aa7fcbf35b88a33f3048df7c3ef66e705510"
    },
    {
        "name": "LahrCard",
        "url": "https://lahrcard.de",
        "ck": "ck_f4f7aabb603c303df851866713acd1583e986919",
        "cs": "cs_fc36e8af3868bffeb9f025b591a50bb5c9fe5164"
    },
    {
        "name": "LandshutCard",
        "url": "https://landshutcard.de",
        "ck": "ck_e57c0904c71eec30ca46ea01b6f0c2fc294b0507",
        "cs": "cs_7c1ead4c323f94b53c4ca5dc838b967ec91f4b2d"
    },
    {
        "name": "LE-Card",
        "url": "https://le-card.de",
        "ck": "ck_66dc8d7ce2f7f91bf681c3ca4aa237ac253b5a81",
        "cs": "cs_0bbcbeae230d5cf726c3568256bc9b597d9f0462"
    },
    {
        "name": "NeubulachCard",
        "url": "https://neubulachcard.de",
        "ck": "ck_c450d4aa2540c7e9e71fdd79f0414c9c18a836a8",
        "cs": "cs_79afae837c26c938f925bf6396ce0163a45b84ce"
    },
    {
        "name": "NeuriedCard",
        "url": "https://neuriedcard.de",
        "ck": "ck_fba02994524dadf212a64ddaede3b80ad785eab0",
        "cs": "cs_821f8b007529330b67017ddc980ea10254ef0ac1"
    },
    {
        "name": "PfulbenCard",
        "url": "https://pfulbencard.de",
        "ck": "ck_8c4bd563d7f42e7914a477c7ca12de18257ebe8a",
        "cs": "cs_e4ebe8ee9a15e2a8e9fd9ef2e1b6d67e85884299"
    },
    {
        "name": "RatingenCard",
        "url": "https://ratingen-card.de",
        "ck": "ck_74a952fe0e6687875d7d11abb41a5420052f6a20",
        "cs": "cs_24df5b3a3f3e14603d828fa3783d7585288dfb56"
    },
    {
        "name": "SRCARD",
        "url": "https://srcard.de",
        "ck": "ck_1f6570848bb7bdc72ba7923b288c240fd57e21ab",
        "cs": "cs_8d827aa1e17a67fe1192aec884ff81daf6df126e"
    },
    {
        "name": "TROCARD",
        "url": "https://trocard.de",
        "ck": "ck_7e0fae280b19b608593bf638b2f9bbaf61160ee6",
        "cs": "cs_7e8a741a7914761e8aee79e8f131da385a68d735"
    },
    {
        "name": "ViertälerCard",
        "url": "https://viertaelercard.de",
        "ck": "ck_14c551d13bc6b53cf317271cf4bd3dbf4436f030",
        "cs": "cs_aa8b02efe281022474b39f7313e87d0c89329a8e"
    },
    {
        "name": "WürmtalCard",
        "url": "https://wuermtalcard.de",
        "ck": "ck_8162c0ca3c6d0d6cfda46ee6fa1383e80b9dcb74",
        "cs": "cs_90aa4c3789d1eec206bef93f43ad1e94389590ed"
    },
    {
        "name": "WunCard",
        "url": "https://wuncard.de",
        "ck": "ck_c4d515eec829afaf701153bcd4ba4d540b19c8f6",
        "cs": "cs_16c7a33dd87fe7628784521adf8f6d4bac55a1aa"
    },
    #  {
    #     "name": "HÜCARD",
    #     "url": "https://huecard.de",
    #     "ck": "ck_1ab4ad8bb7806c0dd70b36b94438ebe204a82e2f",
    #     "cs": "cs_a45791bb6a80f4345df6423673553c9bbdff5ed8"
    # },
        #  {
    #     "name": "MannheimCARD",
    #     "url": "https://ma-card.de",
    #     "ck": "ck_761ce51c295ab2969632edf1f0f11b47534ecfac",
    #     "cs": "cs_e8ad15b3c17363e6142bfdc17c676337ab99afc9"
    # }
]

# ============================
# DATUMSBEREICH DEFINIEREN
# Wer einen benutzerdefinierten Zeitraum exportieren möchte, kann diesen hier festlegen. Einfach auf True setzen und die gewünschten Start- und Enddaten im Format JJJJ-MM-TT eingeben.
# Alternativ kann man auch die zweite Variante nutzen, um alle Bestellungen ab einem bestimmten Startdatum zu exportieren (z.B. Monatsanfang oder 15. des Monats). In diesem Fall einfach USE_CUSTOM_RANGE auf False setzen und ggf. das gewünschte Startdatum anpassen.
# Wichtig: Das Startdatum muss immer mit "T00:00:00" und das Enddatum mit "T23:59:59" ergänzt werden, damit alle Bestellungen des jeweiligen Tages erfasst werden. Wenn kein Enddatum angegeben wird, werden alle Bestellungen ab dem Startdatum exportiert.
# Hinweis: Das Skript exportiert immer alle Bestellungen ab dem Startdatum, auch wenn das Enddatum in der Vergangenheit liegt. Es werden also keine Bestellungen ausgeschlossen, die nach dem Enddatum liegen, solange sie nach dem Startdatum liegen.
# Beispiel für benutzerdefinierten Zeitraum: Start: 2026-03-29, Ende: 2026-03-31 -> Es werden alle Bestellungen vom 29.03.2026 bis einschließlich 31.03.2026 exportiert.
# ============================
today = datetime.now()
USE_CUSTOM_RANGE = False

if USE_CUSTOM_RANGE:
    EXPORT_START = "2026-03-29"
    EXPORT_END = "2026-03-31"

    start_date = f"{EXPORT_START}T00:00:00"
    end_date = f"{EXPORT_END}T23:59:59"
else:
    # 2. Variante: Alle Bestellungen ab einem bestimmten Startdatum exportieren (z.B. Monatsanfang oder 15. des Monats)
    # Einfach das gewünschte Startdatum festlegen (z.B. heute mit Tag 1 oder Tag 15) und ggf. ein Enddatum definieren. Wenn kein Enddatum angegeben wird, werden alle Bestellungen ab dem Startdatum exportiert.
    USE_DAY_15 = False
    today = datetime.now()

    start_date = today.replace(day=15 if USE_DAY_15 else 1).strftime("%Y-%m-%dT00:00:00")
    end_date = None


def map_zweck(value):
    value = str(value).strip().lower()
    if value == "privatkauf":
        return "ja"
    elif value == "firmenkauf":
        return "nein"
    return ""


def to_int(value):
    if value in (None, ""):
        return ""
    try:
        return int(float(str(value).replace(",", ".")))
    except Exception:
        return value


def to_text(value):
    if value in (None, ""):
        return ""
    return str(value).strip()


def safe_str(value):
    if value is None:
        return ""
    return str(value).strip()


def format_date(date_string):
    if not date_string:
        return ""
    try:
        dt = datetime.fromisoformat(date_string.replace("Z", "+00:00"))
        return dt.strftime("%d.%m.%Y %H:%M")
    except Exception:
        return date_string


def format_expiry_date(date_string):
    if not date_string:
        return ""
    try:
        dt = datetime.strptime(date_string, "%Y-%m-%d %H:%M:%S")
        return dt.strftime("%d.%m.%Y %H:%M")
    except Exception:
        return date_string


def get_meta_value(meta_data, key_name):
    for meta in meta_data:
        if meta.get("key") == key_name:
            return meta.get("value")
    return ""


def normalize_voucher_codes(raw_codes):
    if raw_codes is None:
        return []

    if isinstance(raw_codes, list):
        return [str(code).strip() for code in raw_codes if str(code).strip()]

    raw = str(raw_codes).strip()
    if not raw:
        return []

    separators = ["\n", ",", "|", ";"]
    codes = [raw]

    for sep in separators:
        new_codes = []
        for chunk in codes:
            new_codes.extend(chunk.split(sep))
        codes = new_codes

    return [code.strip() for code in codes if code.strip()]


def normalize_price(value):
    if value in (None, ""):
        return ""

    try:
        text = str(value).strip().replace(",", ".")
        num = float(text)

        if "." not in text and "," not in str(value) and num > 999:
            num = num / 100

        return round(num, 2)
    except Exception:
        return value


def extract_price_from_name(name):
    if not name:
        return ""

    match = re.search(r"(\d+(?:[.,]\d+)?)\s*€", name)
    if match:
        return normalize_price(match.group(1))

    return ""


def sanitize_sheet_title(title):
    invalid_chars = ['[', ']', ':', '*', '?', '/', '\\']
    cleaned = title
    for ch in invalid_chars:
        cleaned = cleaned.replace(ch, "_")
    return cleaned[:31] if cleaned else "Shop"


def format_worksheet(ws, fieldnames, currency_columns, integer_columns, header_font, header_fill, header_alignment, data_alignment):
    text_columns = {"Gutschein Code", "PayPal Order ID", "Phone"}

    # Kopfzeile formatieren
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 24

    # Daten formatieren
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            column_name = fieldnames[cell.column - 1]
            cell.alignment = data_alignment

            if column_name in text_columns:
                cell.number_format = "@"
            elif column_name in currency_columns and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
            elif column_name in integer_columns and isinstance(cell.value, (int, float)):
                cell.number_format = '0'

    # Spaltenbreite automatisch
    for col_idx, column_name in enumerate(fieldnames, start=1):
        max_length = len(column_name)

        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue

            if isinstance(value, float) and column_name in currency_columns:
                text = f"{value:,.2f}"
            else:
                text = str(value)

            if len(text) > max_length:
                max_length = len(text)

        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 60)


fieldnames = [
    "Shop",
    "Gutschein Code",
    "Produktname",
    "Preis",
    "Auftraggeber",
    "Zweck",
    "Order ID",
    "Order Date",
    "Payment Method",
    "PayPal Order ID",
    "Order Discount",
    "PayPal Fee",
    "PayPal Net",
    "Menge"
]

currency_columns = {
    "Preis",
    "Order Discount",
    "PayPal Fee",
    "PayPal Net"
}

integer_columns = {
    "Order ID",
    "Menge"
}

shop_rows = {}
all_rows_combined = []

for shop in shops:
    rows = []
    page = 1

    while True:

        params = {
            "per_page": 100,
            "page": page,
            "after": start_date
        }

        if end_date:
            params["before"] = end_date

        response = requests.get(
            f"{shop['url']}/wp-json/wc/v3/orders",
            auth=(shop["ck"], shop["cs"]),
            params=params,
            timeout=30
        )

        response.raise_for_status()
        orders = response.json()

        if not orders:
            break

        for order in orders:
            billing = order.get("billing", {})
            order_meta = order.get("meta_data", [])
            line_items = order.get("line_items", [])

            customer_name = f"{safe_str(billing.get('first_name'))} {safe_str(billing.get('last_name'))}".strip()
            email = safe_str(billing.get("email"))
            phone = safe_str(billing.get("phone"))

            address_line_1 = " ".join([
                safe_str(billing.get("address_1")),
                safe_str(billing.get("address_2"))
            ]).strip()

            address_line_2 = " ".join([
                safe_str(billing.get("city")),
                safe_str(billing.get("state")),
                safe_str(billing.get("country")),
                safe_str(billing.get("postcode"))
            ]).strip()

            order_id = to_int(order.get("id"))
            order_date = format_date(order.get("date_created"))
            payment_method = safe_str(order.get("payment_method_title"))
            order_total = normalize_price(order.get("total"))
            order_discount = normalize_price(order.get("discount_total"))
            status = safe_str(order.get("status"))

            raw_zweck = (
                safe_str(get_meta_value(order_meta, "_billing_options")) or
                safe_str(get_meta_value(order_meta, "billing_options"))
            )
            zweck = map_zweck(raw_zweck)

            paypal_order_id = safe_str(get_meta_value(order_meta, "_ppcp_paypal_order_id"))
            fee_paypal = normalize_price(get_meta_value(order_meta, "_paypal_fee"))
            net_paypal = normalize_price(get_meta_value(order_meta, "_paypal_net"))

            vou_details = get_meta_value(order_meta, "_woo_vou_meta_order_details")
            if not isinstance(vou_details, dict):
                vou_details = {}

            for item in line_items:
                item_id = str(item.get("id", ""))
                item_name = safe_str(item.get("name"))
                item_quantity = to_int(item.get("quantity"))
                item_meta = item.get("meta_data", [])

                if "gutschein pdf" in item_name.lower():
                    auftraggeber = "nein, Auftraggeber"
                else:
                    auftraggeber = "PayPal"

                recipient_email_obj = get_meta_value(item_meta, "_woo_vou_recipient_email")
                recipient_message_obj = get_meta_value(item_meta, "_woo_vou_recipient_message")

                if isinstance(recipient_email_obj, dict):
                    recipient_email = safe_str(recipient_email_obj.get("value"))
                else:
                    recipient_email = safe_str(recipient_email_obj)

                if isinstance(recipient_message_obj, dict):
                    recipient_message = safe_str(recipient_message_obj.get("value"))
                else:
                    recipient_message = safe_str(recipient_message_obj)

                voucher_price = normalize_price(get_meta_value(item_meta, "_woo_vou_voucher_price"))
                if voucher_price == "":
                    voucher_price = extract_price_from_name(item_name)

                raw_voucher_codes = get_meta_value(item_meta, "_woo_vou_codes")
                voucher_codes = normalize_voucher_codes(raw_voucher_codes)
                if not voucher_codes:
                    voucher_codes = [""]

                exp_date = ""
                pdf_template = ""

                if item_id in vou_details and isinstance(vou_details[item_id], dict):
                    exp_date = format_expiry_date(vou_details[item_id].get("exp_date", ""))
                    pdf_template = safe_str(vou_details[item_id].get("pdf_template", ""))

                for voucher_code in voucher_codes:
                    row_data = {
                        "Shop": shop["name"],
                        "Gutschein Code": to_text(voucher_code),
                        "Produktname": item_name,
                        "Preis": voucher_price,
                        "Auftraggeber": auftraggeber,
                        "Zweck": zweck,
                        "Order ID": order_id,
                        "Order Date": order_date,
                        "Payment Method": payment_method,
                        "PayPal Order ID": paypal_order_id,
                        "Order Discount": order_discount,
                        "PayPal Fee": fee_paypal,
                        "PayPal Net": net_paypal,
                        "Menge": item_quantity
                    }

                    rows.append(row_data)
                    all_rows_combined.append(row_data)

        page += 1

    shop_rows[shop["name"]] = rows

wb = Workbook()
default_sheet = wb.active
wb.remove(default_sheet)

header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_alignment = Alignment(vertical="top", wrap_text=True)

# Blatt "Alle Shops"
ws_all = wb.create_sheet(title="Alle Shops")
ws_all.append(fieldnames)
for row in all_rows_combined:
    ws_all.append([row.get(col, "") for col in fieldnames])

format_worksheet(
    ws_all,
    fieldnames,
    currency_columns,
    integer_columns,
    header_font,
    header_fill,
    header_alignment,
    data_alignment
)

# Einzelne Shop-Blätter
for shop_name, rows in shop_rows.items():
    ws = wb.create_sheet(title=sanitize_sheet_title(shop_name))
    ws.append(fieldnames)

    for row in rows:
        ws.append([row.get(col, "") for col in fieldnames])

    format_worksheet(
        ws,
        fieldnames,
        currency_columns,
        integer_columns,
        header_font,
        header_fill,
        header_alignment,
        data_alignment
    )

# Wichtig: Der Dateiname enthält immer das aktuelle Datum,im Master-File bei den XVERWEISEN muss ggf. angepasst werden, damit die Datei gefunden wird. 
# Alternativ kann man auch die zweite Variante nutzen, um den Dateinamen mit Start- und Enddatum zu generieren, wenn ein benutzerdefinierter Zeitraum genutzt wird. In diesem Fall einfach die entsprechende Zeile aktivieren und die andere auskommentieren.
#output_file = f"alle_bestellungen_pro_shop_{today.strftime('%Y-%m-%d')}.xlsx"
# Alternative Dateinamen mit Start- und Enddatum (wenn benutzerdefinierter Zeitraum genutzt wird):
#output_file = f"alle_bestellungen_{EXPORT_START}_bis_{EXPORT_END}.xlsx"
output_file = f"alle_bestellungen_pro_shop.xlsx"

wb.save(output_file)

total_rows = sum(len(rows) for rows in shop_rows.values())
print(f"Excel erstellt: {output_file}")
print(f"Shops: {len(shop_rows)} | Zeilen gesamt: {total_rows} | Startdatum: {start_date}")