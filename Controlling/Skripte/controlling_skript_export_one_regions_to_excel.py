import requests
from datetime import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

shops = [
   
    {
        "name": "ECHT FREIBURG CARD",
        "url": "https://echtfreiburgcard.de",
        "ck": "ck_4442abedf50a5217c4c40a331cf3433c0e797cb5",
        "cs": "cs_c2bfc2d2136ecd6d4d642f5aa3a35fe364912365"
    }
]


# False = ab dem 1. des aktuellen Monats
# True  = ab dem 15. des aktuellen Monats
USE_DAY_15 = False

today = datetime.now()
start_date = today.replace(day=15 if USE_DAY_15 else 1).strftime("%Y-%m-%dT00:00:00")

def map_zweck(value):
    value = str(value).strip().lower()

    if value == "privatkauf":
        return "ja"
    elif value == "firmenkauf":
        return "nein"

    return ""


def to_int(value):
    if value in (None, ""):
        return ""
    try:
        return int(float(str(value).replace(",", ".")))
    except Exception:
        return value


def to_text(value):
    if value in (None, ""):
        return ""
    return str(value).strip()


def safe_str(value):
    if value is None:
        return ""
    return str(value).strip()


def format_date(date_string):
    if not date_string:
        return ""
    try:
        dt = datetime.fromisoformat(date_string.replace("Z", "+00:00"))
        return dt.strftime("%d.%m.%Y %H:%M")
    except Exception:
        return date_string


def format_expiry_date(date_string):
    if not date_string:
        return ""
    try:
        dt = datetime.strptime(date_string, "%Y-%m-%d %H:%M:%S")
        return dt.strftime("%d.%m.%Y %H:%M")
    except Exception:
        return date_string


def get_meta_value(meta_data, key_name):
    for meta in meta_data:
        if meta.get("key") == key_name:
            return meta.get("value")
    return ""


def normalize_voucher_codes(raw_codes):
    if raw_codes is None:
        return []

    if isinstance(raw_codes, list):
        return [str(code).strip() for code in raw_codes if str(code).strip()]

    raw = str(raw_codes).strip()
    if not raw:
        return []

    separators = ["\n", ",", "|", ";"]
    codes = [raw]

    for sep in separators:
        new_codes = []
        for chunk in codes:
            new_codes.extend(chunk.split(sep))
        codes = new_codes

    return [code.strip() for code in codes if code.strip()]


def normalize_price(value):
    """
    Normalisiert Preise:
    - 5000 -> 50.00
    - 5298 -> 52.98
    - 50 -> 50.00
    """
    if value in (None, ""):
        return ""

    try:
        text = str(value).strip().replace(",", ".")
        num = float(text)

        # Heuristik: ganzzahlige Cent-Werte > 999 als Cent interpretieren
        if "." not in text and "," not in str(value) and num > 999:
            num = num / 100

        return round(num, 2)
    except Exception:
        return value


def extract_price_from_name(name):
    """
    Extrahiert z. B. '20€ Gutschein' -> 20
    """
    if not name:
        return ""

    match = re.search(r"(\d+(?:[.,]\d+)?)\s*€", name)
    if match:
        return normalize_price(match.group(1))

    return ""


def sanitize_sheet_title(title):
    """
    Excel-Blattnamen: max 31 Zeichen, keine []:*?/\\
    """
    invalid_chars = ['[', ']', ':', '*', '?', '/', '\\']
    cleaned = title
    for ch in invalid_chars:
        cleaned = cleaned.replace(ch, "_")
    return cleaned[:31] if cleaned else "Shop"


fieldnames = [
    "Shop",
    "Gutschein Code",
    "Produktname",
    "Preis",
    "Auftraggeber",
    #"Name",
    #"Email",
    #"Address 1",
    #"Address 2",
    #"Phone",
    "Zweck",
    "Order ID",
    "Order Date",
    "Payment Method",
    "PayPal Order ID",
    #"Order Total",
    "Order Discount",
    "PayPal Fee",
    "PayPal Net",
    #"Expires",
    #"PDF Template",
    #"E-Mail Adresse des Empfängers",
    #"Nachricht für den Empfänger",
    #"Status",
    "Menge"
]

currency_columns = {
    "Preis",
    "Order Total",
    "Order Discount",
    "PayPal Fee",
    "PayPal Net"
}

integer_columns = {
    "Order ID",
    "Menge"
}

shop_rows = {}

for shop in shops:
    rows = []
    page = 1

    while True:
        response = requests.get(
            f"{shop['url']}/wp-json/wc/v3/orders",
            auth=(shop["ck"], shop["cs"]),
            params={
                "per_page": 100,
                "page": page,
                "after": start_date
            },
            timeout=30
        )

        response.raise_for_status()
        orders = response.json()

        if not orders:
            break

        for order in orders:
            billing = order.get("billing", {})
            order_meta = order.get("meta_data", [])
            line_items = order.get("line_items", [])

            customer_name = f"{safe_str(billing.get('first_name'))} {safe_str(billing.get('last_name'))}".strip()
            email = safe_str(billing.get("email"))
            phone = safe_str(billing.get("phone"))

            address_line_1 = " ".join([
                safe_str(billing.get("address_1")),
                safe_str(billing.get("address_2"))
            ]).strip()

            address_line_2 = " ".join([
                safe_str(billing.get("city")),
                safe_str(billing.get("state")),
                safe_str(billing.get("country")),
                safe_str(billing.get("postcode"))
            ]).strip()

            order_id = safe_str(order.get("id"))
            order_date = format_date(order.get("date_created"))
            payment_method = safe_str(order.get("payment_method_title"))
            order_total = normalize_price(order.get("total"))
            order_discount = normalize_price(order.get("discount_total"))
            status = safe_str(order.get("status"))

            #zweck = (
              #  safe_str(get_meta_value(order_meta, "_billing_options")) or
             #   safe_str(get_meta_value(order_meta, "billing_options"))
            #)

            raw_zweck = (
            safe_str(get_meta_value(order_meta, "_billing_options")) or
            safe_str(get_meta_value(order_meta, "billing_options"))
            )

            zweck = map_zweck(raw_zweck)

            paypal_order_id = safe_str(get_meta_value(order_meta, "_ppcp_paypal_order_id"))
            fee_paypal = normalize_price(get_meta_value(order_meta, "_paypal_fee"))
            net_paypal = normalize_price(get_meta_value(order_meta, "_paypal_net"))

            vou_details = get_meta_value(order_meta, "_woo_vou_meta_order_details")
            if not isinstance(vou_details, dict):
                vou_details = {}

            for item in line_items:
                item_id = str(item.get("id", ""))
                item_name = safe_str(item.get("name"))
                item_quantity = safe_str(item.get("quantity"))
                item_meta = item.get("meta_data", [])

                if "gutschein pdf" in item_name.lower():
                    auftraggeber = "nein, Auftraggeber"
                else:
                    auftraggeber = "PayPal"

                recipient_email_obj = get_meta_value(item_meta, "_woo_vou_recipient_email")
                recipient_message_obj = get_meta_value(item_meta, "_woo_vou_recipient_message")

                if isinstance(recipient_email_obj, dict):
                    recipient_email = safe_str(recipient_email_obj.get("value"))
                else:
                    recipient_email = safe_str(recipient_email_obj)

                if isinstance(recipient_message_obj, dict):
                    recipient_message = safe_str(recipient_message_obj.get("value"))
                else:
                    recipient_message = safe_str(recipient_message_obj)

                voucher_price = normalize_price(get_meta_value(item_meta, "_woo_vou_voucher_price"))
                if voucher_price == "":
                    voucher_price = extract_price_from_name(item_name)

                raw_voucher_codes = get_meta_value(item_meta, "_woo_vou_codes")
                voucher_codes = normalize_voucher_codes(raw_voucher_codes)
                if not voucher_codes:
                    voucher_codes = [""]

                exp_date = ""
                pdf_template = ""

                if item_id in vou_details and isinstance(vou_details[item_id], dict):
                    exp_date = format_expiry_date(vou_details[item_id].get("exp_date", ""))
                    pdf_template = safe_str(vou_details[item_id].get("pdf_template", ""))

                for voucher_code in voucher_codes:
                    rows.append({
                        "Shop": shop["name"],
                        "Gutschein Code": to_text(voucher_code),
                        "Produktname": item_name,
                        "Preis": voucher_price,
                        "Auftraggeber": auftraggeber,
                        "Name": customer_name,
                        "Email": email,
                        "Address 1": address_line_1,
                        "Address 2": address_line_2,
                        "Phone": phone,
                        "Zweck": zweck,
                        "Order ID": to_int(order_id),
                        "Order Date": order_date,
                        "Payment Method": payment_method,
                        "PayPal Order ID": paypal_order_id,
                        "Order Total": order_total,
                        "Order Discount": order_discount,
                        "PayPal Fee": fee_paypal,
                        "PayPal Net": net_paypal,
                        "Expires": exp_date,
                        "PDF Template": pdf_template,
                        "E-Mail Adresse des Empfängers": recipient_email,
                        "Nachricht für den Empfänger": recipient_message,
                        "Status": status,
                        "Menge": to_int(item_quantity)
                    })

        page += 1

    shop_rows[shop["name"]] = rows

wb = Workbook()
default_sheet = wb.active
wb.remove(default_sheet)

header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_alignment = Alignment(vertical="top", wrap_text=True)

for shop_name, rows in shop_rows.items():
    ws = wb.create_sheet(title=sanitize_sheet_title(shop_name))

    # Kopfzeile
    ws.append(fieldnames)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Daten
    for row in rows:
        ws.append([row.get(col, "") for col in fieldnames])

    # Freeze Pane
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Zahlenformatierung und Alignment
    text_columns = {"Gutschein Code", "PayPal Order ID", "Phone"}

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            column_name = fieldnames[cell.column - 1]
            cell.alignment = data_alignment

            if column_name in text_columns:
                cell.number_format = "@"
            elif column_name in currency_columns and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
            elif column_name in integer_columns and isinstance(cell.value, (int, float)):
                cell.number_format = '0'

    # Spaltenbreite automatisch
    for col_idx, column_name in enumerate(fieldnames, start=1):
        max_length = len(column_name)

        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue

            if isinstance(value, float) and column_name in currency_columns:
                text = f"{value:,.2f}"
            else:
                text = str(value)

            if len(text) > max_length:
                max_length = len(text)

        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Zeilenhöhe Header
    ws.row_dimensions[1].height = 24

output_file = f"alle_bestellungen_einer_region.xlsx"
#output_file = f"alle_bestellungen_einer_region_{today.strftime('%Y-%m-%d')}.xlsx"
wb.save(output_file)

total_rows = sum(len(rows) for rows in shop_rows.values())
print(f"Excel erstellt: {output_file}")
print(f"Shops: {len(shop_rows)} | Zeilen gesamt: {total_rows} | Startdatum: {start_date}")