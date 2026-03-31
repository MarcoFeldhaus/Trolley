import requests
import json
import os
import re
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, BarChart, Reference

from config import shops, OUTPUT_FILE, STATE_FILE

# =========================
# HELPERS
# =========================

def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")


def safe_str(v):
    return "" if v is None else str(v).strip()


def to_int(v):
    try:
        return int(float(str(v).replace(",", ".")))
    except:
        return ""


def format_date(date_string):
    try:
        return datetime.fromisoformat(date_string.replace("Z", "+00:00")).strftime("%d.%m.%Y %H:%M")
    except:
        return ""


def normalize_price(v):
    try:
        return round(float(str(v).replace(",", ".")), 2)
    except:
        return ""


def extract_price_from_name(name):
    match = re.search(r"(\d+(?:[.,]\d+)?)\s*€", name or "")
    return normalize_price(match.group(1)) if match else ""


def get_meta_value(meta_data, key):
    for m in meta_data:
        if m.get("key") == key:
            return m.get("value")
    return ""


def normalize_voucher_codes(raw):
    if not raw:
        return []
    if isinstance(raw, list):
        return [str(x).strip() for x in raw if str(x).strip()]
    raw = str(raw)
    for sep in ["\n", ",", ";", "|"]:
        raw = raw.replace(sep, ",")
    return [x.strip() for x in raw.split(",") if x.strip()]


def map_zweck(v):
    v = safe_str(v).lower()
    if v == "privatkauf":
        return "ja"
    if v == "firmenkauf":
        return "nein"
    return ""


# =========================
# STATE
# =========================

def load_last_run():
    if os.path.exists(STATE_FILE):
        return json.load(open(STATE_FILE)).get("last_run")
    return None


def save_last_run():
    json.dump({"last_run": datetime.utcnow().isoformat()}, open(STATE_FILE, "w"))


# =========================
# API
# =========================

def fetch_orders(shop, start_date):
    page = 1
    orders = []

    while True:
        try:
            url = f"{shop['url']}/wp-json/wc/v3/orders"
            log(f"API-Aufruf: {url} | after={start_date} | page={page}")
            
            r = requests.get(
                url,
                auth=(shop["ck"], shop["cs"]),
                params={"per_page": 100, "page": page, "after": start_date},
                timeout=30
            )
            r.raise_for_status()
            data = r.json()

            if not data:
                break

            orders.extend(data)
            page += 1

        except requests.exceptions.HTTPError as e:
            log(f"HTTP-Fehler {shop['name']}: {e.response.status_code} - {e.response.text}")
            break
        except Exception as e:
            log(f"Fehler {shop['name']}: {e}")
            break

    return orders


# =========================
# CORE LOGIC
# =========================

def extract_voucher_codes(item_meta, order_meta, item_id):
    # 1. item_meta
    raw = get_meta_value(item_meta, "_woo_vou_codes")

    # 2. fallback: order_meta → vou_details
    if not raw:
        vou_details = get_meta_value(order_meta, "_woo_vou_meta_order_details")

        if isinstance(vou_details, str):
            try:
                vou_details = json.loads(vou_details)
            except:
                vou_details = {}

        if isinstance(vou_details, dict) and item_id in vou_details:
            raw = vou_details[item_id].get("codes")

    codes = normalize_voucher_codes(raw)
    return codes if codes else [""]


def transform_orders(orders, shop_name, existing_keys):
    rows = []

    for order in orders:
        order_id = to_int(order.get("id"))
        order_date = format_date(order.get("date_created"))
        order_meta = order.get("meta_data", [])

        raw_zweck = (
            get_meta_value(order_meta, "_billing_options") or
            get_meta_value(order_meta, "billing_options")
        )
        zweck = map_zweck(raw_zweck)

        for item in order.get("line_items", []):
            item_id = str(item.get("id"))
            name = safe_str(item.get("name"))
            qty = to_int(item.get("quantity"))
            item_meta = item.get("meta_data", [])

            price = normalize_price(get_meta_value(item_meta, "_woo_vou_voucher_price"))
            if not price:
                price = extract_price_from_name(name)

            codes = extract_voucher_codes(item_meta, order_meta, item_id)

            for code in codes:
                key = (order_id, code)

                # 🔥 Duplikat vermeiden
                if key in existing_keys:
                    continue
                # Auftraggeber ermitteln
                if "gutschein pdf" in name.lower():
                    auftraggeber = "nein, Auftraggeber"
                else:
                    auftraggeber = "PayPal"

                row = {
                    "Shop": shop_name,
                    "Gutschein Code": code,
                    "Order ID": order_id,
                    "Produkt": name,
                    "Datum": order_date,
                    "Preis": price,
                    "Zweck": zweck,
                    "Auftraggeber": auftraggeber,
                    "Menge": qty
                }

                rows.append(row)
                existing_keys.add(key)

    return rows

# =========================
# LOAD EXISTING EXCEL
# =========================

def load_existing():
    if not os.path.exists(OUTPUT_FILE):
        return [], set()
    wb = load_workbook(OUTPUT_FILE)
    if "Alle Shops" not in wb.sheetnames:
        return [], set()
    ws = wb["Alle Shops"]
    headers = [c.value for c in ws[1]]
    rows = []
    keys = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, r))
        rows.append(d)
        order_id = d.get("Order ID")
        voucher_code = d.get("Gutschein Code")
        if order_id and voucher_code:
            keys.add((order_id, voucher_code))
    return rows, keys

# =========================
# BUILD MONTHLY REPORT
# =========================

def build_monthly_report(rows):
    report = defaultdict(lambda: {"umsatz": 0, "bestellungen": set()})
    for r in rows:
        datum = r.get("Datum")
        if not datum:
            continue
        try:
            month = str(datum)[:7]
            key = (r.get("Shop"), month)
            preis = float(r.get("Preis") or 0)
            report[key]["umsatz"] += preis
            report[key]["bestellungen"].add(r.get("Order ID"))
        except (ValueError, TypeError):
            continue
    
    final = []
    for (shop, month), data in report.items():
        final.append({
            "Shop": shop,
            "Monat": month,
            "Umsatz": round(data["umsatz"], 2),
            "Bestellungen": len(data["bestellungen"])
        })
    return final

# =========================
# DASHBOARD + KPI
# =========================

def add_dashboard(wb, monthly_rows, all_rows):
    ws = wb.create_sheet("Dashboard")

    # Aggregationen
    revenue_per_month = defaultdict(float)
    revenue_per_shop = defaultdict(float)
    orders_per_month = defaultdict(int)
    for r in monthly_rows:
        revenue_per_month[r["Monat"]] += r["Umsatz"]
        orders_per_month[r["Monat"]] += r["Bestellungen"]
        revenue_per_shop[r["Shop"]] += r["Umsatz"]

    # KPI
    total_revenue = sum(revenue_per_month.values())
    total_orders = sum(orders_per_month.values())
    avg_order = total_revenue / total_orders if total_orders else 0

    ws["A1"] = "KPI"
    ws["A2"] = "Umsatz"
    ws["B2"] = total_revenue
    ws["A3"] = "Bestellungen"
    ws["B3"] = total_orders
    ws["A4"] = "Ø Bestellwert"
    ws["B4"] = round(avg_order, 2)
    ws["B2"].number_format = '#,##0.00'
    ws["B4"].number_format = '#,##0.00'

    # Umsatz pro Monat
    ws["A6"], ws["B6"] = "Monat", "Umsatz"
    row = 7
    for m in sorted(revenue_per_month):
        ws.cell(row=row, column=1, value=m)
        ws.cell(row=row, column=2, value=revenue_per_month[m])
        row += 1
    chart = LineChart()
    data = Reference(ws, min_col=2, min_row=6, max_row=row-1)
    cats = Reference(ws, min_col=1, min_row=7, max_row=row-1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.title = "Umsatz pro Monat"
    chart.style = 10
    ws.add_chart(chart, "D6")

    # Wachstum (MoM)
    ws["D6"], ws["E6"] = "Wachstum Monat %", ""
    prev = None
    row_growth = 7
    for m in sorted(revenue_per_month):
        current = revenue_per_month[m]
        growth = ((current - prev) / prev * 100) if prev else 0
        ws.cell(row=row_growth, column=4, value=growth)
        ws.cell(row=row_growth, column=4).number_format = '0.00%'
        row_growth += 1
        prev = current

    # Top 5 Produkte
    product_sales = defaultdict(float)
    
    for r in all_rows:
        produkt = r.get("Produkt") or "Unbekannt"
        try:
            preis = float(r.get("Preis") or 0)
            menge = float(r.get("Menge") or 0)
            product_sales[produkt] += preis * menge
        except (ValueError, TypeError):
            continue
    

    top5 = sorted(product_sales.items(), key=lambda x: x[1], reverse=True)[:5]
    ws["A20"], ws["B20"] = "Top Produkte", "Umsatz"
    r_top = 21
    for name, value in top5:
        ws.cell(row=r_top, column=1, value=name)
        ws.cell(row=r_top, column=2, value=round(value, 2))
        ws.cell(row=r_top, column=2).number_format = '#,##0.00'
        r_top += 1

    # Umsatz pro Shop
    start_shop = r_top + 1
    ws.cell(row=start_shop, column=1, value="Shop")
    ws.cell(row=start_shop, column=2, value="Umsatz")
    r_shop = start_shop + 1
    for s, v in revenue_per_shop.items():
        ws.cell(row=r_shop, column=1, value=s)
        ws.cell(row=r_shop, column=2, value=v)
        ws.cell(row=r_shop, column=2).number_format = '#,##0.00'
        r_shop += 1
    chart2 = BarChart()
    data2 = Reference(ws, min_col=2, min_row=start_shop, max_row=r_shop-1)
    cats2 = Reference(ws, min_col=1, min_row=start_shop+1, max_row=r_shop-1)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.title = "Umsatz pro Shop"
    chart2.style = 10
    ws.add_chart(chart2, "D20")

# =========================
# WRITE EXCEL
# =========================

def write_excel(all_rows, monthly_rows):
    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "Alle Shops"
    headers = ["Shop", "Gutschein Code", "Order ID", "Produktname", "Datum", "Preis", "Privatkauf", "Auftraggeber", "Menge"]
    ws_all.append(headers)
    for r in all_rows:
        ws_all.append([r.get(h) for h in headers])
    
    # Separate Blatt pro Shop
    shops_set = sorted({r["Shop"] for r in all_rows})
    for shop in shops_set:
        ws = wb.create_sheet(shop)
        ws.append(headers)
        for r in all_rows:
            if r["Shop"] == shop:
                ws.append([r.get(h) for h in headers])
    
    # Monatsreport
    ws_month = wb.create_sheet("Monatsreport")
    headers_m = ["Shop", "Monat", "Umsatz", "Bestellungen"]
    ws_month.append(headers_m)
    for r in sorted(monthly_rows, key=lambda x: (x["Monat"], x["Shop"])):
        ws_month.append([r.get(h) for h in headers_m])
    
    # Dashboard
    add_dashboard(wb, monthly_rows, all_rows)
    
    wb.save(OUTPUT_FILE)

def write_pdf(all_rows, monthly_rows, output_file):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        
        pdf_output = output_file.replace(".xlsx", ".pdf")
        doc = SimpleDocTemplate(pdf_output, pagesize=landscape(A4), topMargin=0.5*cm, bottomMargin=0.5*cm)
        
        # Daten vorbereiten
        headers = ["Shop", "Gutschein Code", "Order ID", "Produktname", "Datum", "Preis", "Privatkauf", "Auftraggeber", "Menge"]
        data = [headers]
        
        for row in all_rows:
            data.append([str(row.get(h, "")) for h in headers])
        
        # Tabelle mit Formatting
        table = Table(data, colWidths=[1.6*cm]*len(headers))
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')])
        ]))
        elements = [table]
        doc.build(elements) 
        log(f"PDF erfolgreich erstellt: {pdf_output}")
    except ImportError:
        log("ReportLab nicht installiert. PDF wird nicht erstellt.")    



# =========================
# MAIN
# =========================

def main():
    log("Start")

    start_date = load_last_run() or datetime.now().replace(day=1).isoformat()
    existing_rows, existing_keys = load_existing()
    all_rows = existing_rows[:]
    
    for shop in shops:
        log(f"Lade {shop['name']}")

        orders = fetch_orders(shop, start_date)
        rows = transform_orders(orders, shop["name"], existing_keys)

        log(f"{shop['name']}: {len(rows)} neue Zeilen")
        all_rows.extend(rows)

    monthly_rows = build_monthly_report(all_rows)
    write_excel(all_rows, monthly_rows)
    write_pdf(all_rows, monthly_rows, OUTPUT_FILE)  # ← Diese Zeile hinzufügen
    save_last_run()

    log(f"Fertig: {len(all_rows)} Zeilen")


if __name__ == "__main__":
    main()