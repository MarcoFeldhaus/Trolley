import requests
import json
import os
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.styles import Font

from config import shops, OUTPUT_FILE, STATE_FILE
from config import shops
print("Anzahl Shops:", len(shops))
for s in shops:
    print(s["name"])
# =========================
# STATE HANDLING
# =========================

def load_last_run():
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, "r") as f:
            return json.load(f).get("last_run")
    return None

def save_last_run():
    json.dump({"last_run": datetime.utcnow().isoformat()}, open(STATE_FILE, "w"))

# =========================
# API FETCH
# =========================

def fetch_orders(shop, start_date):
    page = 1
    all_orders = []

    while True:
        params = {"per_page": 100, "page": page, "after": start_date}
        response = requests.get(
            f"{shop['url']}/wp-json/wc/v3/orders",
            auth=(shop["ck"], shop["cs"]),
            params=params,
            timeout=30
        )
        response.raise_for_status()
        data = response.json()
        if not data:
            break
        all_orders.extend(data)
        page += 1
    return all_orders

# =========================
# TRANSFORM
# =========================

def transform_orders(orders, shop_name):
    rows = []
    for order in orders:
        order_id = order.get("id")
        date = order.get("date_created", "")[:10]
        for item in order.get("line_items", []):
            rows.append({
                "Shop": shop_name,
                "Order ID": order_id,
                "Datum": date,
                "Produkt": item.get("name"),
                "Menge": item.get("quantity"),
                "Preis": round(float(item.get("price", 0)), 2)
            })
    return rows

# =========================
# LOAD EXISTING
# =========================

def load_existing():
    if not os.path.exists(OUTPUT_FILE):
        return [], set()
    wb = load_workbook(OUTPUT_FILE)
    if "Alle Shops" not in wb.sheetnames:
        return [], set()
    ws = wb["Alle Shops"]
    headers = [c.value for c in ws[1]]
    rows = []
    keys = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, r))
        rows.append(d)
        keys.add((d["Order ID"], d["Produkt"]))
    return rows, keys

# =========================
# DASHBOARD DATA
# =========================

def build_dashboard_data(monthly_rows):
    revenue_per_month = {}
    orders_per_month = {}
    revenue_per_shop = {}
    for r in monthly_rows:
        m = r["Monat"]
        shop = r["Shop"]
        revenue_per_month[m] = revenue_per_month.get(m, 0) + r["Umsatz"]
        orders_per_month[m] = orders_per_month.get(m, 0) + r["Bestellungen"]
        revenue_per_shop[shop] = revenue_per_shop.get(shop, 0) + r["Umsatz"]
    return revenue_per_month, orders_per_month, revenue_per_shop

def calculate_growth(monthly_rows):
    revenue_per_month = {}
    for r in monthly_rows:
        revenue_per_month[r["Monat"]] = revenue_per_month.get(r["Monat"], 0) + r["Umsatz"]
    months_sorted = sorted(revenue_per_month.keys())
    growth_data = []
    prev = None
    for m in months_sorted:
        current = revenue_per_month[m]
        growth = 0 if prev is None else ((current - prev) / prev * 100 if prev != 0 else 0)
        growth_data.append({"Monat": m, "Umsatz": current, "Wachstum %": round(growth, 2)})
        prev = current
    return growth_data

def get_top_products(all_rows):
    product_sales = {}
    for r in all_rows:
        name = r.get("Produkt")
        revenue = float(r.get("Preis", 0)) * int(r.get("Menge", 1))
        product_sales[name] = product_sales.get(name, 0) + revenue
    top = sorted(product_sales.items(), key=lambda x: x[1], reverse=True)[:5]
    return top

# =========================
# DASHBOARD
# =========================

def add_dashboard(wb, monthly_rows, all_rows):
    ws = wb.create_sheet("Dashboard")
    revenue_m, orders_m, revenue_s = build_dashboard_data(monthly_rows)

    # KPI
    total_revenue = sum(revenue_m.values())
    total_orders = sum(orders_m.values())
    avg_order = total_revenue / total_orders if total_orders else 0

    ws["A1"] = "KPI"
    ws["A2"] = "Umsatz"
    ws["B2"] = total_revenue
    ws["A3"] = "Bestellungen"
    ws["B3"] = total_orders
    ws["A4"] = "Ø Bestellwert"
    ws["B4"] = round(avg_order, 2)

    ws["B2"].number_format = '#,##0.00'
    ws["B4"].number_format = '#,##0.00'

    # Umsatz pro Monat
    ws["A6"] = "Monat"
    ws["B6"] = "Umsatz"
    row = 7
    for m, v in sorted(revenue_m.items()):
        ws.cell(row=row, column=1, value=m)
        ws.cell(row=row, column=2, value=v)
        row += 1

    chart = LineChart()
    data = Reference(ws, min_col=2, min_row=6, max_row=row-1)
    cats = Reference(ws, min_col=1, min_row=7, max_row=row-1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.title = "Umsatz pro Monat"
    chart.style = 10
    ws.add_chart(chart, "D6")

    # Wachstum
    growth_data = calculate_growth(monthly_rows)
    start_growth = row + 2
    ws.cell(row=start_growth, column=1, value="Monat")
    ws.cell(row=start_growth, column=2, value="Umsatz")
    ws.cell(row=start_growth, column=3, value="Wachstum %")

    r = start_growth + 1
    for g in growth_data:
        ws.cell(row=r, column=1, value=g["Monat"])
        ws.cell(row=r, column=2, value=g["Umsatz"])
        ws.cell(row=r, column=3, value=g["Wachstum %"])
        # Farbe
        if g["Wachstum %"] > 0:
            ws.cell(row=r, column=3).font = Font(color="008000")
        elif g["Wachstum %"] < 0:
            ws.cell(row=r, column=3).font = Font(color="FF0000")
        r += 1

    # Top 5 Produkte
    top_products = get_top_products(all_rows)
    start_top = r + 2
    ws.cell(row=start_top, column=1, value="Top Produkte")
    ws.cell(row=start_top + 1, column=1, value="Produkt")
    ws.cell(row=start_top + 1, column=2, value="Umsatz")
    r_top = start_top + 2
    for name, value in top_products:
        ws.cell(row=r_top, column=1, value=name)
        ws.cell(row=r_top, column=2, value=round(value, 2))
        ws.cell(row=r_top, column=2).number_format = '#,##0.00'
        r_top += 1

    # Umsatz pro Shop
    start_shop = r_top + 2
    ws.cell(row=start_shop, column=1, value="Shop")
    ws.cell(row=start_shop, column=2, value="Umsatz")
    r_shop = start_shop + 1
    for s, v in revenue_s.items():
        ws.cell(row=r_shop, column=1, value=s)
        ws.cell(row=r_shop, column=2, value=v)
        ws.cell(row=r_shop, column=2).number_format = '#,##0.00'
        r_shop += 1

    chart2 = BarChart()
    data2 = Reference(ws, min_col=2, min_row=start_shop, max_row=r_shop-1)
    cats2 = Reference(ws, min_col=1, min_row=start_shop+1, max_row=r_shop-1)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.title = "Umsatz pro Shop"
    chart2.style = 10
    ws.add_chart(chart2, "D20")

# =========================
# MONATSREPORT
# =========================

def build_monthly_report(rows):
    report = defaultdict(lambda: {"umsatz": 0, "bestellungen": set()})
    for r in rows:
        if not r.get("Datum"):
            continue
        month = r["Datum"][:7]  # YYYY-MM
        key = (r["Shop"], month)
        report[key]["umsatz"] += float(r.get("Preis", 0))
        report[key]["bestellungen"].add(r["Order ID"])

    final = []
    for (shop, month), data in report.items():
        final.append({
            "Shop": shop,
            "Monat": month,
            "Umsatz": round(data["umsatz"], 2),
            "Bestellungen": len(data["bestellungen"])
        })
    return final

# =========================
# WRITE EXCEL
# =========================

def write_excel(all_rows, monthly_rows):
    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "Alle Shops"
    headers = ["Shop", "Order ID", "Datum", "Produkt", "Menge", "Preis"]
    ws_all.append(headers)
    for r in all_rows:
        ws_all.append([r.get(h) for h in headers])

    add_dashboard(wb, monthly_rows, all_rows)

    ws_month = wb.create_sheet("Monatsreport")
    headers_m = ["Shop", "Monat", "Umsatz", "Bestellungen"]
    ws_month.append(headers_m)
    for r in sorted(monthly_rows, key=lambda x: (x["Monat"], x["Shop"])):
        ws_month.append([r.get(h) for h in headers_m])

    wb.save(OUTPUT_FILE)

# =========================
# MAIN
# =========================

def main():
    last_run = load_last_run()
    start_date = last_run if last_run else datetime.now().replace(day=1).isoformat()

    existing_rows, existing_keys = load_existing()
    all_rows = existing_rows.copy()

    for shop in shops:
        orders = fetch_orders(shop, start_date)
        rows = transform_orders(orders, shop["name"])
        for r in rows:
            key = (r["Order ID"], r["Produkt"])
            if key not in existing_keys:
                all_rows.append(r)

    monthly = build_monthly_report(all_rows)
    write_excel(all_rows, monthly)
    save_last_run()
    print(f"Neue Zeilen: {len(all_rows) - len(existing_rows)}")
    print("Fertig!")

if __name__ == "__main__":
    main()