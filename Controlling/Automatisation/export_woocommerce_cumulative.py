import requests
from datetime import datetime
import re
import json
import os
import threading
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict
from openpyxl.chart import LineChart, BarChart, Reference

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

from config import shops, OUTPUT_FILE, STATE_FILE

# ============================
# HILFSFUNKTIONEN
# ============================

def log(msg):
    """Gibt eine Nachricht mit Zeitstempel auf der Konsole aus."""
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")


def input_with_timeout(prompt, timeout_seconds=5, default_value="1"):
    """Wartet auf Eingabe mit Timeout. Nach Ablauf wird der Standard-Wert verwendet."""
    result = [None]
    
    def input_thread():
        result[0] = input(prompt)
    
    thread = threading.Thread(target=input_thread, daemon=True)
    thread.start()
    thread.join(timeout=timeout_seconds)
    
    if result[0] is None:
        print(f"\n(Keine Eingabe nach {timeout_seconds} Sekunden - verwende Standard: {default_value})")
        return default_value
    return result[0]


def map_zweck(value):
    """Mappt Zweck-Wert: 'privatkauf' -> 'ja', 'firmenkauf' -> 'nein'."""
    value = str(value).strip().lower()
    if value in ["firmenkauf", "t", ""]:
        return "nein"
    elif value == "privatkauf":
        return "ja"
    return ""


def to_int(value):
    """Konvertiert einen Wert robust zu Integer."""
    if value in (None, ""):
        return ""
    try:
        return int(float(str(value).replace(",", ".")))
    except Exception:
        return value


def to_text(value):
    """Konvertiert einen Wert zu Text."""
    if value in (None, ""):
        return ""
    return str(value).strip()


def safe_str(value):
    """Sichere String-Konvertierung."""
    if value is None:
        return ""
    return str(value).strip()


def format_date(date_string):
    """Formatiert ISO-Datum in deutsches Datumsformat."""
    if not date_string:
        return ""
    try:
        dt = datetime.fromisoformat(date_string.replace("Z", "+00:00"))
        return dt.strftime("%d.%m.%Y")
    except Exception:
        return date_string


def normalize_price(value):
    """Konvertiert Preis zu Float mit 2 Dezimalstellen."""
    if value in (None, ""):
        return ""

    try:
        text = str(value).strip().replace(",", ".")
        num = float(text)

        if "." not in text and "," not in str(value) and num > 999:
            num = num / 100

        return round(num, 2)
    except Exception:
        return value


def extract_price_from_name(name):
    """Extrahiert Preis aus Produktnamen."""
    if not name:
        return ""

    match = re.search(r"(\d+(?:[.,]\d+)?)\s*€", name)
    if match:
        return normalize_price(match.group(1))

    return ""


def get_meta_value(meta_data, key_name):
    """Sucht einen Wert in WooCommerce Meta-Daten."""
    for meta in meta_data:
        if meta.get("key") == key_name:
            return meta.get("value")
    return ""


def normalize_voucher_codes(raw_codes):
    """Normalisiert Gutscheincodes aus verschiedenen Formaten. Nur 15-stellige Integer-Codes werden akzeptiert."""
    if raw_codes is None:
        return []

    codes = []
    
    if isinstance(raw_codes, list):
        for code in raw_codes:
            code_str = str(code).strip()
            if len(code_str) == 15 and code_str.isdigit():
                codes.append(int(code_str))
    else:
        raw = str(raw_codes).strip()
        if raw:
            separators = ["\n", ",", "|", ";"]
            code_list = [raw]
            
            for sep in separators:
                new_codes = []
                for chunk in code_list:
                    new_codes.extend(chunk.split(sep))
                code_list = new_codes
            
            for code_str in code_list:
                code_str = code_str.strip()
                if len(code_str) == 15 and code_str.isdigit():
                    codes.append(int(code_str))
    
    return codes


def sanitize_sheet_title(title):
    """Bereinigt Blattnamen für Excel."""
    invalid_chars = ['[', ']', ':', '*', '?', '/', '\\']
    cleaned = title
    for ch in invalid_chars:
        cleaned = cleaned.replace(ch, "_")
    return cleaned[:31] if cleaned else "Shop"


def add_pdf_export(all_rows, shop_rows, output_file):
    """Exportiert Dashboard auf Seite 1 + pro Region eine Seite mit Tabellendaten."""
    if not HAS_REPORTLAB:
        log("Warnung: reportlab nicht installiert - PDF-Export übersprungen")
        return
    
    try:
        pdf_file = output_file.replace(".xlsx", ".pdf")
        
        # Aggregiere Dashboard-Daten
        total_revenue = sum(row.get("Preis", 0) or 0 for row in all_rows if isinstance(row.get("Preis"), (int, float)))
        total_vouchers = len(all_rows)
        
        regional_data = defaultdict(lambda: {"count": 0, "revenue": 0})
        for row in all_rows:
            shop = row.get("Shop", "")
            preis = row.get("Preis", 0) or 0
            if shop:
                regional_data[shop]["count"] += 1
                if isinstance(preis, (int, float)):
                    regional_data[shop]["revenue"] += preis
        
        # PDF-Inhalt erstellen
        pdf = SimpleDocTemplate(pdf_file, pagesize=landscape(A4), rightMargin=0.5*cm, leftMargin=0.5*cm)
        elements = []
        styles = getSampleStyleSheet()
        
        # Styles definieren
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('0x1F4E78'),
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        shop_title_style = ParagraphStyle(
            'ShopTitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('0x1F4E78'),
            spaceAfter=15,
            fontName='Helvetica-Bold'
        )
        
        # ===== SEITE 1: DASHBOARD =====
        elements.append(Paragraph("EXPORT REPORT - DASHBOARD", title_style))
        elements.append(Spacer(1, 0.3*cm))
        
        # KPIs
        kpi_data = [
            ["GESAMTUMSATZ", f"€ {total_revenue:.2f}"],
            ["GESAMTE GUTSCHEINE", str(total_vouchers)]
        ]
        kpi_table = Table(kpi_data, colWidths=[8*cm, 6*cm])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('0xADD8E6')),
            ('BACKGROUND', (1, 0), (1, -1), colors.HexColor('0xE8F4F8')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        elements.append(kpi_table)
        elements.append(Spacer(1, 0.5*cm))
        
        # Regional Übersicht
        elements.append(Paragraph("Pro Region:", shop_title_style))
        
        regional_table_data = [["Shop", "Anzahl Gutscheine", "Umsatz"]]
        for shop in sorted(regional_data.keys()):
            data = regional_data[shop]
            regional_table_data.append([
                shop,
                str(data["count"]),
                f"€ {data['revenue']:.2f}"
            ])
        
        regional_table = Table(regional_table_data, colWidths=[9*cm, 5*cm, 5*cm])
        regional_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('0x1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('0xF0F0F0')])
        ]))
        elements.append(regional_table)
        
        # ===== PRO REGION: EINE SEITE MIT DATEN =====
        for shop_name in sorted(shop_rows.keys()):
            rows = shop_rows[shop_name]
            if not rows:
                continue
            
            elements.append(PageBreak())
            elements.append(Paragraph(f"REGION: {shop_name}", shop_title_style))
            elements.append(Spacer(1, 0.3*cm))
            
            # Datentabelle pro Shop
            shop_table_data = [["Gutschein Code", "Produktname", "Preis", "Datum", "Privatkauf"]]
            
            for row in rows:
                code = row.get("Gutschein Code", "")
                code_str = str(code) if code else ""
                
                shop_table_data.append([
                    code_str,
                    row.get("Produktname", "")[:30],
                    f"€ {row.get('Preis', 0) or 0:.2f}",
                    row.get("Datum", ""),
                    row.get("Privatkauf", "")
                ])
            
            shop_table = Table(shop_table_data, colWidths=[5*cm, 6*cm, 3*cm, 3.5*cm, 3*cm])
            shop_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('0x1F4E78')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('0xF5F5F5')])
            ]))
            elements.append(shop_table)
        
        # PDF bauen
        pdf.build(elements)
        log(f"PDF mit Regionen exportiert: {pdf_file}")
    
    except Exception as e:
        log(f"Fehler beim PDF-Export: {e}")


def add_dashboard(wb, all_rows):
    """Erstellt ein Dashboard mit KPIs und 2 großen Charts."""
    if not all_rows:
        return
    
    ws = wb.create_sheet(title="Dashboard", index=0)
    
    # KPI Daten berechnen
    total_revenue = sum(row.get("Preis", 0) or 0 for row in all_rows if isinstance(row.get("Preis"), (int, float)))
    total_vouchers = len(all_rows)
    
    # Regional aggregieren
    regional_data = defaultdict(lambda: {"count": 0, "revenue": 0})
    monthly_data = defaultdict(lambda: {"count": 0, "revenue": 0})
    monthly_regional_data = defaultdict(lambda: defaultdict(lambda: {"count": 0, "revenue": 0}))
    
    for row in all_rows:
        shop = row.get("Shop", "")
        datum = row.get("Datum", "")
        preis = row.get("Preis", 0) or 0
        
        if shop:
            regional_data[shop]["count"] += 1
            if isinstance(preis, (int, float)):
                regional_data[shop]["revenue"] += preis
        
        if datum:
            try:
                month_str = ".".join(datum.split(".")[-2:])
                monthly_data[month_str]["count"] += 1
                if isinstance(preis, (int, float)):
                    monthly_data[month_str]["revenue"] += preis
                
                # Pro Monat + Region
                if shop:
                    monthly_regional_data[month_str][shop]["count"] += 1
                    if isinstance(preis, (int, float)):
                        monthly_regional_data[month_str][shop]["revenue"] += preis
            except:
                pass
    
    # Dashboard Header
    ws['A1'] = "DASHBOARD"
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    # KPIs
    ws['A3'] = "Gesamtumsatz"
    ws['B3'] = total_revenue
    ws['B3'].number_format = '#,##0.00 €'
    
    for cell in ['A3']:
        ws[cell].font = Font(bold=True)
        ws[cell].fill = PatternFill(fill_type="solid", fgColor="ADD8E6")
    
    # Pro Monat Header (mit Regions-Details)
    ws['A6'] = "Pro Monat (mit Region)"
    ws['A6'].font = Font(bold=True, size=11)
    ws['A7'] = "Monat"
    ws['B7'] = "Region"
    ws['C7'] = "Anzahl Gutscheine"
    ws['D7'] = "Umsatz"
    
    for cell in ['A7', 'B7', 'C7', 'D7']:
        ws[cell].font = Font(bold=True, color="FFFFFF")
        ws[cell].fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        ws[cell].alignment = Alignment(horizontal="center")
    
    # Monatliche Daten pro Region
    row_idx = 8
    for month in sorted(monthly_regional_data.keys()):
        for shop in sorted(monthly_regional_data[month].keys()):
            data = monthly_regional_data[month][shop]
            ws[f'A{row_idx}'] = month
            ws[f'B{row_idx}'] = shop
            ws[f'C{row_idx}'] = data["count"]
            ws[f'D{row_idx}'] = round(data["revenue"], 2)
            ws[f'D{row_idx}'].number_format = '#,##0.00'
            row_idx += 1
    
    # Regional Report Header
    ws[f'A{row_idx + 2}'] = "Pro Region (Gesamt)"
    ws[f'A{row_idx + 2}'].font = Font(bold=True, size=11)
    
    header_row = row_idx + 3
    ws[f'A{header_row}'] = "Shop"
    ws[f'B{header_row}'] = "Anzahl Gutscheine"
    ws[f'C{header_row}'] = "Umsatz"
    
    for cell in [f'A{header_row}', f'B{header_row}', f'C{header_row}']:
        ws[cell].font = Font(bold=True, color="FFFFFF")
        ws[cell].fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        ws[cell].alignment = Alignment(horizontal="center")
    
    # Shop-Daten
    shop_row = header_row + 1
    for shop in sorted(regional_data.keys()):
        data = regional_data[shop]
        ws[f'A{shop_row}'] = shop
        ws[f'B{shop_row}'] = data["count"]
        ws[f'C{shop_row}'] = round(data["revenue"], 2)
        ws[f'C{shop_row}'].number_format = '#,##0.00'
        shop_row += 1
    
    # Chart-Hilfsspalten ganz nach unten verschieben (ab Zeile nach Shop-Daten)
    chart_data_start = shop_row + 2
    row_idx = shop_row
    
    # Spaltenbreiten
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    
    # === CHARTS AB F2 ===
    if row_idx > 8:
        # Chart 1: Wachstum per Monat (LineChart) - ab F2, 22cm x 14cm
        chart_growth = LineChart()
        chart_growth.title = "Wachstum pro Monat"
        chart_growth.style = 10
        chart_growth.x_axis.title = "Monat"
        chart_growth.y_axis.title = "Anzahl Gutscheine"
        chart_growth.height = 14
        chart_growth.width = 22
        
        # Monatsdaten in Hilfsspalten (K:L) schreiben (nach den Tabellen)
        ws[f'K{chart_data_start}'] = "Monat"
        ws[f'L{chart_data_start}'] = "Anzahl"
        ws[f'K{chart_data_start}'].font = Font(bold=True)
        ws[f'L{chart_data_start}'].font = Font(bold=True)
        
        month_row = chart_data_start + 1
        for month in sorted(monthly_data.keys()):
            data = monthly_data[month]
            ws[f'K{month_row}'] = month
            ws[f'L{month_row}'] = data["count"]
            month_row += 1
        
        data_growth = Reference(ws, min_col=12, min_row=chart_data_start, max_row=month_row-1)
        cats_growth = Reference(ws, min_col=11, min_row=chart_data_start+1, max_row=month_row-1)
        chart_growth.add_data(data_growth, titles_from_data=True)
        chart_growth.set_categories(cats_growth)
        
        ws.add_chart(chart_growth, "F2")
        
        # Chart 2: Umsatz pro Shop (BarChart) - ab F23, gleiche Größe, untereinander
        chart_revenue = BarChart()
        chart_revenue.type = "col"
        chart_revenue.title = "Umsatz pro Shop"
        chart_revenue.style = 10
        chart_revenue.x_axis.title = "Shop"
        chart_revenue.y_axis.title = "Umsatz (€)"
        chart_revenue.height = 14
        chart_revenue.width = 22
        
        # Shop-Umsatzdaten in Hilfsspalten (M:N) schreiben
        ws[f'M{chart_data_start}'] = "Shop"
        ws[f'N{chart_data_start}'] = "Umsatz"
        ws[f'M{chart_data_start}'].font = Font(bold=True)
        ws[f'N{chart_data_start}'].font = Font(bold=True)
        
        shop_row_chart = chart_data_start + 1
        for shop in sorted(regional_data.keys()):
            data = regional_data[shop]
            ws[f'M{shop_row_chart}'] = shop
            ws[f'N{shop_row_chart}'] = round(data["revenue"], 2)
            shop_row_chart += 1
        
        data_revenue = Reference(ws, min_col=14, min_row=chart_data_start, max_row=shop_row_chart-1)
        cats_revenue = Reference(ws, min_col=13, min_row=chart_data_start+1, max_row=shop_row_chart-1)
        chart_revenue.add_data(data_revenue, titles_from_data=True)
        chart_revenue.set_categories(cats_revenue)
        
        ws.add_chart(chart_revenue, "F37")
    
    ws.print_area = f'A1:C{row_idx}'


def format_worksheet(ws, fieldnames, currency_columns, integer_columns, header_font, header_fill, header_alignment, data_alignment):
    """Formatiert Worksheet mit Header, Breiten und Zahlenformaten."""
    text_columns = {"PayPal Order ID", "Phone"}

    # Kopfzeile formatieren
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 24

    # Daten formatieren
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            column_name = fieldnames[cell.column - 1]
            cell.alignment = data_alignment

            if column_name in text_columns:
                cell.number_format = "@"
            elif column_name in currency_columns and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
            elif column_name in integer_columns:
                if cell.value is None or cell.value == "":
                    cell.value = ""
                elif isinstance(cell.value, int):
                    cell.number_format = '0'

    # Spaltenbreite automatisch
    for col_idx, column_name in enumerate(fieldnames, start=1):
        max_length = len(column_name)

        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue

            if isinstance(value, float) and column_name in currency_columns:
                text = f"{value:,.2f}"
            else:
                text = str(value)

            if len(text) > max_length:
                max_length = len(text)

        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 60)


# ============================
# STATE MANAGEMENT
# ============================

def load_last_run():
    """Lädt den letzten Lauf aus state.json."""
    if os.path.exists(STATE_FILE):
        try:
            return json.load(open(STATE_FILE)).get("last_run")
        except:
            return None
    return None


def save_last_run():
    """Speichert den aktuellen Zeitstempel in state.json."""
    json.dump({"last_run": datetime.now().isoformat()}, open(STATE_FILE, "w"))


def load_existing_data(output_file):
    """Lädt existierende Daten aus der Excel-Datei (falls vorhanden)."""
    if not os.path.exists(output_file):
        return [], {}
    
    try:
        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        
        # Laden aus "Alle Shops" Blatt
        existing_rows = []
        if "Alle Shops" in wb.sheetnames:
            ws = wb["Alle Shops"]
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] is None:  # Skip empty rows
                    continue
                row_data = {
                    "Shop": row[0],
                    "Gutschein Code": row[1],
                    "Produktname": row[2],
                    "Preis": row[3],
                    "Auftraggeber": row[4],
                    "Privatkauf": row[5],
                    "Order ID": row[6],
                    "Datum": row[7]
                }
                existing_rows.append(row_data)
        
        # Laden aus einzelnen Shop-Blättern für shop_rows
        existing_shop_rows = {}
        for sheet_name in wb.sheetnames:
            if sheet_name not in ["Alle Shops", "Dashboard"]:
                ws = wb[sheet_name]
                shop_data = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is None:  # Skip empty rows
                        continue
                    row_data = {
                        "Shop": row[0],
                        "Gutschein Code": row[1],
                        "Produktname": row[2],
                        "Preis": row[3],
                        "Auftraggeber": row[4],
                        "Privatkauf": row[5],
                        "Order ID": row[6],
                        "Datum": row[7]
                    }
                    shop_data.append(row_data)
                existing_shop_rows[sheet_name] = shop_data
        
        log(f"Existierende Daten geladen: {len(existing_rows)} Zeilen")
        return existing_rows, existing_shop_rows
    
    except Exception as e:
        log(f"Fehler beim Laden existierender Daten: {e}")
        return [], {}


def merge_rows(existing_rows, new_rows):
    """Mergt alte und neue Rows. Duplikate (by Gutschein Code) werden vermieden."""
    # Dictionary mit existierenden Codes
    existing_codes = {row["Gutschein Code"]: row for row in existing_rows if row["Gutschein Code"]}
    
    # Kombinierte Liste starten mit existierenden Daten
    combined = list(existing_rows)
    
    # Neue Daten hinzufügen, falls Gutschein-Code nicht bereits existiert
    for row in new_rows:
        if row["Gutschein Code"] not in existing_codes:
            combined.append(row)
            existing_codes[row["Gutschein Code"]] = row
    
    log(f"Daten gemergt: {len(existing_rows)} existierend + {len(new_rows)} neu = {len(combined)} total (Duplikate vermieden)")
    return combined


def get_date_range():
    """Ermöglicht Auswahl zwischen State-Modus oder manuellen Start/End-Daten."""
    print("\n=== EXPORT-MODUS ===")
    print("1 = State verwenden (inkrementeller Export seit letztem Lauf)")
    print("2 = Start- & End-Datum eingeben (Zeitraum exportieren)")
    choice = input_with_timeout("Wähle Modus (1 oder 2): ", timeout_seconds=5, default_value="1").strip()
    
    if choice == "2":
        while True:
            try:
                start_input = input_with_timeout("Start-Datum (DD.MM.YYYY): ", timeout_seconds=15, default_value="").strip()
                end_input = input_with_timeout("End-Datum (DD.MM.YYYY): ", timeout_seconds=15, default_value="").strip()
                
                start_dt = datetime.strptime(start_input, "%d.%m.%Y")
                end_dt = datetime.strptime(end_input, "%d.%m.%Y")
                
                start_date = f"{start_dt.strftime('%Y-%m-%d')}T00:00:00"
                end_date = f"{end_dt.strftime('%Y-%m-%d')}T23:59:59"
                
                start_str = start_dt.strftime("%d%m%Y")
                end_str = end_dt.strftime("%d%m%Y")
                output_file = OUTPUT_FILE.replace(".xlsx", f"_{start_str}_{end_str}.xlsx")
                
                log(f"Export-Zeitraum: {start_input} bis {end_input}")
                log(f"Output-Datei: {output_file}")
                return start_date, end_date, output_file, False
            except ValueError:
                print("Fehler: Ungültiges Datumsformat. Bitte DD.MM.YYYY verwenden.")
    else:
        state_date = load_last_run()
        if not state_date:
            # Fallback: Letzten Monat ab 1. exportieren
            now = datetime.now()
            start_date = now.replace(day=1).isoformat()
        else:
            start_date = state_date
        log(f"Export-Modus: State | Start: {start_date}")
        return start_date, None, OUTPUT_FILE, True


# ============================
# HAUPTLOGIK
# ============================

fieldnames = [
    "Shop",
    "Gutschein Code",
    "Produktname",
    "Preis",
    "Auftraggeber",
    "Privatkauf",
    "Order ID",
    "Datum"
]

currency_columns = {"Preis"}
integer_columns = {"Order ID", "Gutschein Code"}

# Lade existierende Daten (kumulativ sammeln)
start_date, end_date, output_file, use_state = get_date_range()
existing_all_rows, existing_shop_rows = load_existing_data(output_file)

shop_rows = dict(existing_shop_rows)  # Starte mit existierenden Shop-Daten
all_rows_combined = list(existing_all_rows)  # Starte mit existierenden Daten
new_rows_combined = []  # Sammle nur die neuen Daten für Merging

for shop in shops:
    new_rows = []  # Nur neue Daten aus dieser API-Iteration
    page = 1

    while True:
        params = {
            "per_page": 100,
            "page": page,
            "after": start_date
        }

        if end_date:
            params["before"] = end_date

        try:
            log(f"API-Aufruf: {shop['url']}/wp-json/wc/v3/orders | page={page}")
            response = requests.get(
                f"{shop['url']}/wp-json/wc/v3/orders",
                auth=(shop["ck"], shop["cs"]),
                params=params,
                timeout=30
            )
            response.raise_for_status()
            orders = response.json()

            if not orders:
                break

            for order in orders:
                order_meta = order.get("meta_data", [])
                line_items = order.get("line_items", [])

                order_id = to_int(order.get("id"))
                order_date = format_date(order.get("date_created"))

                raw_zweck = (
                    safe_str(get_meta_value(order_meta, "_billing_options")) or
                    safe_str(get_meta_value(order_meta, "billing_options"))
                )
                zweck = map_zweck(raw_zweck)

                for item in line_items:
                    item_id = str(item.get("id", ""))
                    item_name = safe_str(item.get("name"))
                    item_quantity = to_int(item.get("quantity"))
                    item_meta = item.get("meta_data", [])

                    if "gutschein pdf" in item_name.lower():
                        auftraggeber = "nein, Auftraggeber"
                    else:
                        auftraggeber = "PayPal"

                    voucher_price = normalize_price(get_meta_value(item_meta, "_woo_vou_voucher_price"))
                    if voucher_price == "":
                        voucher_price = extract_price_from_name(item_name)

                    raw_voucher_codes = get_meta_value(item_meta, "_woo_vou_codes")
                    voucher_codes = normalize_voucher_codes(raw_voucher_codes)
                    if not voucher_codes:
                        continue  # Überspringe dieses Item wenn kein gültiger 15-stelliger Gutscheincode vorhanden

                    for voucher_code in voucher_codes:
                        row_data = {
                            "Shop": shop["name"],
                            "Gutschein Code": voucher_code,
                            "Produktname": item_name,
                            "Preis": voucher_price,
                            "Auftraggeber": auftraggeber,
                            "Privatkauf": zweck,
                            "Order ID": order_id,
                            "Datum": order_date
                        }

                        new_rows.append(row_data)
                        new_rows_combined.append(row_data)

            page += 1

        except Exception as e:
            log(f"Fehler bei {shop['name']}: {e}")
            break

    # Merge neue Daten mit existierenden pro Shop
    if shop["name"] in shop_rows:
        shop_rows[shop["name"]] = merge_rows(shop_rows[shop["name"]], new_rows)
    else:
        shop_rows[shop["name"]] = new_rows
    
    log(f"{shop['name']}: {len(new_rows)} neue Zeilen (gesamt: {len(shop_rows[shop['name']])})")

# Merge neue Daten mit existierenden für "Alle Shops"
all_rows_combined = merge_rows(existing_all_rows, new_rows_combined)

# ============================
# EXCEL GENERIEREN
# ============================

wb = Workbook()
default_sheet = wb.active
wb.remove(default_sheet)

header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_alignment = Alignment(vertical="top", wrap_text=True)

# Blatt "Alle Shops"
ws_all = wb.create_sheet(title="Alle Shops")
ws_all.append(fieldnames)
for row in all_rows_combined:
    ws_all.append([row.get(col, "") for col in fieldnames])

format_worksheet(
    ws_all,
    fieldnames,
    currency_columns,
    integer_columns,
    header_font,
    header_fill,
    header_alignment,
    data_alignment
)

# Einzelne Shop-Blätter
for shop_name, rows in shop_rows.items():
    ws = wb.create_sheet(title=sanitize_sheet_title(shop_name))
    ws.append(fieldnames)

    for row in rows:
        ws.append([row.get(col, "") for col in fieldnames])

    format_worksheet(
        ws,
        fieldnames,
        currency_columns,
        integer_columns,
        header_font,
        header_fill,
        header_alignment,
        data_alignment
    )

# Füge Dashboard hinzu
add_dashboard(wb, all_rows_combined)

wb.save(output_file)

# Exportiere PDF
add_pdf_export(all_rows_combined, shop_rows, output_file)

# Speichere State nur wenn im State-Modus
if use_state:
    save_last_run()

total_rows = sum(len(rows) for rows in shop_rows.values())
log(f"Excel erstellt: {output_file}")
log(f"Shops: {len(shop_rows)} | Zeilen gesamt: {total_rows}")
