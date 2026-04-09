import requests
import json
import os
import re
import traceback
from datetime import datetime, timezone
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import PatternFill, Font

from config import shops, OUTPUT_FILE, STATE_FILE

# =========================
# HELPERS
# =========================

def log(msg):
    """Gibt eine Nachricht mit Zeitstempel auf der Konsole aus."""
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")


def safe_str(v):
    """Konvertiert einen Wert sicher zu String, gibt leeren String für None zurück."""
    return "" if v is None else str(v).strip()


def to_int(v):
    """Konvertiert einen Wert robust zu Integer, ignoriert Kommas und Fehler."""
    try:
        return int(float(str(v).replace(",", ".")))
    except:
        return ""


def format_date(date_string):
    """Formatiert ISO-Datum in deutsches Datumsformat DD.MM.YYYY."""
    try:
        return datetime.fromisoformat(date_string.replace("Z", "+00:00")).strftime("%d.%m.%Y")
    except:
        return ""


def normalize_price(v):
    """Konvertiert Preis zu Float mit 2 Dezimalstellen, ersetzt Kommas durch Punkte."""
    try:
        return round(float(str(v).replace(",", ".")), 2)
    except:
        return ""


def extract_price_from_name(name):
    """Extrahiert Preis aus Produktnamen mit €-Symbol (z.B. '20€ Gutschein')."""
    match = re.search(r"(\d+(?:[.,]\d+)?)\s*€", name or "")
    return normalize_price(match.group(1)) if match else ""


def get_meta_value(meta_data, key):
    """Sucht einen Wert in WooCommerce Meta-Daten nach Key, gibt leeren String zurück wenn nicht gefunden."""
    for m in meta_data:
        if m.get("key") == key:
            return m.get("value")
    return ""


def normalize_voucher_codes(raw):
    """Normalisiert Gutscheincodes aus verschiedenen Formaten (String, Liste, mehrere Separatoren)."""
        return []
    if isinstance(raw, list):
        return [str(x).strip() for x in raw if str(x).strip()]
    raw = str(raw)
    for sep in ["\n", ",", ";", "|"]:
        raw = raw.replace(sep, ",")
    return [x.strip() for x in raw.split(",") if x.strip()]


def map_zweck(v):
    """Mappt Zweck-Wert: 'privatkauf' -> 'ja', 'firmenkauf' -> 'nein', sonst leerer String."""
    v = safe_str(v).lower()
    if v == "privatkauf":
        return "ja"
    if v == "firmenkauf":
        return "nein"
    return ""


# =========================
# STATE
# =========================

def load_last_run():
    """Lädt den Zeitstempel des letzten erfolgreichen Skriptlaufs aus state.json."""
    if os.path.exists(STATE_FILE):
        return json.load(open(STATE_FILE)).get("last_run")
    return None


def save_last_run():
    """Speichert den aktuellen UTC-Zeitstempel als letzten erfolgreichen Lauf in state.json."""
    json.dump({"last_run": datetime.now(timezone.utc).isoformat()}, open(STATE_FILE, "w"))


# =========================
# API
# =========================

def fetch_orders(shop, start_date):
    """Ruft paginierte Bestellungen von WooCommerce API ab, startet nach start_date mit Fehlerhandling."""
    page = 1
    orders = []

    while True:
        try:
            url = f"{shop['url']}/wp-json/wc/v3/orders"
            log(f"API-Aufruf: {url} | after={start_date} | page={page}")
            
            r = requests.get(
                url,
                auth=(shop["ck"], shop["cs"]),
                params={"per_page": 100, "page": page, "after": start_date},
                timeout=30
            )
            r.raise_for_status()
            data = r.json()

            if not data:
                break

            orders.extend(data)
            page += 1

        except requests.exceptions.HTTPError as e:
            log(f"HTTP-Fehler {shop['name']}: {e.response.status_code} - {e.response.text}")
            break
        except Exception as e:
            log(f"Fehler {shop['name']}: {e}")
            break

    return orders


# =========================
# CORE LOGIC
# =========================

def extract_voucher_codes(item_meta, order_meta, item_id):
    """Extrahiert Gutscheincodes aus Item- oder Order-Meta-Daten mit Fallback-Logik."""
    # 1. item_meta
    raw = get_meta_value(item_meta, "_woo_vou_codes")

    # 2. fallback: order_meta → vou_details
    if not raw:
        vou_details = get_meta_value(order_meta, "_woo_vou_meta_order_details")

        if isinstance(vou_details, str):
            try:
                vou_details = json.loads(vou_details)
            except:
                vou_details = {}

        if isinstance(vou_details, dict) and item_id in vou_details:
            raw = vou_details[item_id].get("codes")

    codes = normalize_voucher_codes(raw)
    return codes if codes else [""]


def transform_orders(orders, shop_name, existing_keys):
    """Transformiert WooCommerce-Bestellungen in strukturierte Daten mit Duplikatsprüfung und Filterung."""
    rows = []

    for order in orders:
        order_id = to_int(order.get("id"))
        order_date = format_date(order.get("date_created"))
        order_meta = order.get("meta_data", [])

        raw_zweck = (
            get_meta_value(order_meta, "_billing_options") or
            get_meta_value(order_meta, "billing_options")
        )
        zweck = map_zweck(raw_zweck)

        for item in order.get("line_items", []):
            item_id = str(item.get("id"))
            name = safe_str(item.get("name"))
            qty = to_int(item.get("quantity"))
            item_meta = item.get("meta_data", [])

            price = normalize_price(get_meta_value(item_meta, "_woo_vou_voucher_price"))
            if not price:
                price = extract_price_from_name(name)

            codes = extract_voucher_codes(item_meta, order_meta, item_id)
            
            # Nur 14-stellige Gutscheinnummern akzeptieren (keine Testkäufe)
            codes = [c for c in codes if len(str(c).strip()) == 15 and str(c).strip().isdigit()]
            
            # Wenn kein gültiger Code, überspringen
            if not codes:
                continue

            for code in codes:
                key = (order_id, code)

                # 🔥 Duplikat vermeiden
                if key in existing_keys:
                    continue
                # Auftraggeber ermitteln
                if "gutschein pdf" in name.lower():
                    auftraggeber = "nein, Auftraggeber"
                else:
                    auftraggeber = "PayPal"

                row = {
                    "Shop": shop_name,
                    "Gutschein Code": code,
                    "Order ID": order_id,
                    "Produkt": name,
                    "Datum": order_date,
                    "Preis": price,
                    "Zweck": zweck,
                    "Auftraggeber": auftraggeber,
                    "Menge": qty
                }

                rows.append(row)
                existing_keys.add(key)

    return rows

# =========================
# LOAD EXISTING EXCEL
# =========================

def load_existing():
    """Lädt bestehende Daten aus Excel-Datei und bereinigt Datumsformat, ignoriert doppelte Einträge."""
    if not os.path.exists(OUTPUT_FILE):
        return [], set()
    wb = load_workbook(OUTPUT_FILE)
    if "Alle Shops" not in wb.sheetnames:
        return [], set()
    ws = wb["Alle Shops"]
    headers = [c.value for c in ws[1]]
    rows = []
    keys = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, r))
        # Datum korrigieren: entferne Uhrzeit falls vorhanden
        if "Datum" in d and d["Datum"]:
            datum_str = str(d["Datum"])
            if " " in datum_str:
                d["Datum"] = datum_str.split(" ")[0]  # Nur Datum ohne Uhrzeit
        rows.append(d)
        order_id = d.get("Order ID")
        voucher_code = d.get("Gutschein Code")
        if order_id and voucher_code:
            keys.add((order_id, voucher_code))
    return rows, keys

# =========================
# BUILD MONTHLY REPORT
# =========================

def build_monthly_report(rows):
    """Erstellt Monatsbericht mit aggregiertem Umsatz und Bestellungsanzahl pro Shop und Monat."""
    report = defaultdict(lambda: {"umsatz": 0, "bestellungen": set()})
    for r in rows:
        datum = r.get("Datum")
        if not datum:
            continue
        try:
            # Format: "01.04.2026 12:34" -> "2026-04"
            parts = str(datum).split(" ")[0].split(".")
            month = f"{parts[2]}-{parts[1]}" if len(parts) >= 3 else ""
            if not month:
                continue
            key = (r.get("Shop"), month)
            preis = float(r.get("Preis") or 0)
            report[key]["umsatz"] += preis
            report[key]["bestellungen"].add(r.get("Order ID"))
        except (ValueError, TypeError):
            continue
    
    final = []
    for (shop, month), data in report.items():
        final.append({
            "Shop": shop,
            "Monat": month,
            "Umsatz": round(data["umsatz"], 2),
            "Bestellungen": len(data["bestellungen"])
        })
    return final

# =========================
# DASHBOARD + KPI
# =========================

def add_dashboard(wb, monthly_rows, all_rows):
    """Erstellt Dashboard mit KPIs, Wachstumschart und Umsatz pro Shop mit Formatierung."""
    ws = wb.create_sheet("Dashboard")

    # Aggregationen
    revenue_per_month = defaultdict(float)
    revenue_per_shop = defaultdict(float)
    vouchers_per_shop = defaultdict(int)
    orders_per_month = defaultdict(int)
    for r in monthly_rows:
        revenue_per_month[r["Monat"]] += r["Umsatz"]
        orders_per_month[r["Monat"]] += r["Bestellungen"]
        revenue_per_shop[r["Shop"]] += r["Umsatz"]
    
    # Gutscheine pro Shop zählen
    for r in all_rows:
        shop = r.get("Shop")
        if shop:
            vouchers_per_shop[shop] += 1

    # KPI
    total_revenue = sum(revenue_per_month.values())
    total_orders = sum(orders_per_month.values())
    avg_order = total_revenue / total_orders if total_orders else 0

    ws["A1"] = "KPI"
    format_headers(ws, ["KPI"], start_row=1)
    ws["A2"] = "Umsatz"
    ws["B2"] = total_revenue
    ws["A3"] = "Bestellungen"
    ws["B3"] = total_orders
    ws["A4"] = "Ø Bestellwert"
    ws["B4"] = round(avg_order, 2)
    ws["B2"].number_format = '#,##0.00'
    ws["B4"].number_format = '#,##0.00'

    # Umsatz pro Monat
    ws["A6"], ws["B6"] = "Monat", "Umsatz"
    ws["D6"], ws["E6"] = "Wachstum Monat %", ""
    format_headers(ws, ["Monat", "Umsatz"])
    
    # Formatierung für Wachstum Spalte
    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    bold_black_font = Font(bold=True, color="000000", size=11)
    ws["D6"].fill = light_blue_fill
    ws["D6"].font = bold_black_font
    ws.column_dimensions["D"].width = max(len("Wachstum Monat %") + 2, 15)
    
    row = 7
    for m in sorted(revenue_per_month):
        ws.cell(row=row, column=1, value=m)
        ws.cell(row=row, column=2, value=revenue_per_month[m])
        row += 1

    # Wachstum (MoM)
    prev = None
    row_growth = 7
    for m in sorted(revenue_per_month):
        current = revenue_per_month[m]
        growth = ((current - prev) / prev * 100) if prev else 0
        ws.cell(row=row_growth, column=4, value=growth)
        ws.cell(row=row_growth, column=4).number_format = '0.00%'
        row_growth += 1
        prev = current
    
    # Wachstum Chart
    chart_growth = LineChart()
    data_growth = Reference(ws, min_col=4, min_row=6, max_row=row_growth-1)
    cats_growth = Reference(ws, min_col=1, min_row=7, max_row=row_growth-1)
    chart_growth.add_data(data_growth, titles_from_data=True)
    chart_growth.set_categories(cats_growth)
    chart_growth.title = "Wachstum pro Monat"
    chart_growth.style = 10
    chart_growth.width = 22
    chart_growth.height = 14
    ws.add_chart(chart_growth, "K6")

    # Top 5 Produkte
    product_sales = defaultdict(float)
    
    for r in all_rows:
        produkt = r.get("Produkt") or "Unbekannt"
        try:
            preis = float(r.get("Preis") or 0)
            menge = float(r.get("Menge") or 0)
            product_sales[produkt] += preis * menge
        except (ValueError, TypeError):
            continue
    

    top5 = sorted(product_sales.items(), key=lambda x: x[1], reverse=True)[:5]
    ws["A20"], ws["B20"] = "Top Produkte", "Umsatz"
    format_headers(ws, ["Top Produkte", "Umsatz"], start_row=20)
    r_top = 21
    for name, value in top5:
        ws.cell(row=r_top, column=1, value=name)
        ws.cell(row=r_top, column=2, value=round(value, 2))
        ws.cell(row=r_top, column=2).number_format = '#,##0.00'
        r_top += 1

    # Umsatz pro Shop
    start_shop = r_top + 1
    ws.cell(row=start_shop, column=1, value="Region/Shop")
    ws.cell(row=start_shop, column=2, value="Anzahl Gutscheine")
    ws.cell(row=start_shop, column=3, value="Umsatz")
    format_headers(ws, ["Region/Shop", "Anzahl Gutscheine", "Umsatz"], start_row=start_shop)
    r_shop = start_shop + 1
    for s in sorted(revenue_per_shop.keys()):
        ws.cell(row=r_shop, column=1, value=s)
        ws.cell(row=r_shop, column=2, value=vouchers_per_shop[s])
        ws.cell(row=r_shop, column=3, value=revenue_per_shop[s])
        ws.cell(row=r_shop, column=3).number_format = '#,##0.00'
        r_shop += 1
    chart2 = BarChart()
    data2 = Reference(ws, min_col=3, min_row=start_shop, max_row=r_shop-1)
    cats2 = Reference(ws, min_col=1, min_row=start_shop+1, max_row=r_shop-1)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.title = "Umsatz pro Shop"
    chart2.style = 10
    chart2.width = 22
    chart2.height = 14
    ws.add_chart(chart2, "K31")

# =========================
# WRITE EXCEL
# =========================

def format_headers(ws, headers, start_row=1):
    """Formatiert Spaltenüberschriften: hellblauer Hintergrund, fette schwarze Schrift (Größe 11), Auto-Breite."""
    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    bold_black_font = Font(bold=True, color="000000", size=11)
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.fill = light_blue_fill
        cell.font = bold_black_font
        
        # Spaltenbreite basierend auf Headerlänge
        ws.column_dimensions[cell.column_letter].width = max(len(str(header)) + 2, 15)

def write_excel(all_rows, monthly_rows):
    """Schreibt Workbook mit Blättern: Alle Shops, pro Shop einzeln, Monatsreport und Dashboard."""
    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "Alle Shops"
    headers = ["Shop", "Gutschein Code", "Order ID", "Produktname", "Datum", "Preis", "Privatkauf", "Auftraggeber", "Menge"]
    ws_all.append(headers)
    format_headers(ws_all, headers)
    for r in all_rows:
        ws_all.append([r.get(h) for h in headers])
    
    # Separate Blatt pro Shop
    shops_set = sorted({r["Shop"] for r in all_rows})
    for shop in shops_set:
        ws = wb.create_sheet(shop)
        ws.append(headers)
        format_headers(ws, headers)
        for r in all_rows:
            if r["Shop"] == shop:
                ws.append([r.get(h) for h in headers])
    
    # Monatsreport
    ws_month = wb.create_sheet("Monatsreport")
    headers_m = ["Shop", "Monat", "Umsatz", "Bestellungen"]
    ws_month.append(headers_m)
    format_headers(ws_month, headers_m)
    for r in sorted(monthly_rows, key=lambda x: (x["Monat"], x["Shop"])):
        ws_month.append([r.get(h) for h in headers_m])
    
    # Dashboard
    add_dashboard(wb, monthly_rows, all_rows)
    
    wb.save(OUTPUT_FILE)

def write_pdf(all_rows, monthly_rows, output_file):
    """Erstellt optional PDF-Report aus all_rows im Querformat mit Tabellen-Styling (ReportLab erforderlich)."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        
        pdf_output = output_file.replace(".xlsx", ".pdf")
        doc = SimpleDocTemplate(pdf_output, pagesize=landscape(A4), topMargin=0.5*cm, bottomMargin=0.5*cm)
        
        # Daten vorbereiten
        headers = ["Shop", "Gutschein Code", "Order ID", "Produktname", "Datum", "Preis", "Privatkauf", "Auftraggeber", "Menge"]
        data = [headers]
        
        for row in all_rows:
            data.append([str(row.get(h, "")) for h in headers])
        
        # Tabelle mit Formatting
        table = Table(data, colWidths=[1.6*cm]*len(headers))
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')])
        ]))
        elements = [table]
        doc.build(elements) 
        log(f"PDF erfolgreich erstellt: {pdf_output}")
    except ImportError as e:
        log(f"ReportLab nicht installiert: {e}")
    except Exception as e:
        log(f"Fehler beim PDF erstellen: {type(e).__name__}: {e}")
        traceback.print_exc()    



# =========================
# MAIN
# =========================

def main():
    """Hauptfunktion: Lädt letzte Daten, ruft Bestellungen ab, verarbeitet sie, erstellt Excel/PDF und speichert Status."""
    log("Start")

    start_date = load_last_run() or datetime.now().replace(day=1).isoformat()
    existing_rows, existing_keys = load_existing()
    all_rows = existing_rows[:]
    
    for shop in shops:
        log(f"Lade {shop['name']}")

        orders = fetch_orders(shop, start_date)
        rows = transform_orders(orders, shop["name"], existing_keys)

        log(f"{shop['name']}: {len(rows)} neue Zeilen")
        all_rows.extend(rows)

    monthly_rows = build_monthly_report(all_rows)
    write_excel(all_rows, monthly_rows)
    write_pdf(all_rows, monthly_rows, OUTPUT_FILE)  # ← Diese Zeile hinzufügen
    save_last_run()

    log(f"Fertig: {len(all_rows)} Zeilen")


if __name__ == "__main__":
    main()